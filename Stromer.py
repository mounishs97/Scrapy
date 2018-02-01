from openpyxl import load_workbook,Workbook
from selenium import webdriver
import threading
import math
class XlsxVar:
    wb,ws,filename,sheetname = None,None,None,None;
    @staticmethod
    def create():
        try:
            XlsxVar.wb = load_workbook(XlsxVar.filename);
            print("Loading the file ",XlsxVar.filename);
            try:
              XlsxVar.ws = XlsxVar.wb[XlsxVar.sheetname];
              print("Loading the sheet ",XlsxVar.sheetname," in a file ",XlsxVar.filename);
            except KeyError:
               XlsxVar.ws = XlsxVar.wb.create_sheet(XlsxVar.sheetname);
               print("Creating the sheet ",XlsxVar.sheetname);
        except FileNotFoundError:
           XlsxVar.wb = Workbook();
           XlsxVar.ws = XlsxVar.wb.create_sheet(XlsxVar.sheetname,0);
           print("Creating the new file");
    @staticmethod
    def save():
        print("Saving the file");
        XlsxVar.wb.save(XlsxVar.filename);
    @staticmethod
    def close():
        print("Closing the file");
        XlsxVar.wb.close();
    @staticmethod
    def writeHeading():
        XlsxVar.create();
        XlsxVar.ws["A1"] = "Exam No";
        j=2;
        for content in TornodoFetch.new_header:
           XlsxVar.ws.cell(row=1,column=j).value=content;
           j=j+1;
        XlsxVar.save();
        XlsxVar.create();
class TornodoFetch:
    driver,mark,head = None,[],[];
    new_header=[];
    def __init__(self):
        self.driver = webdriver.PhantomJS();
    def URL(self,url):
        self.driver.get(url);
    def insertData(self,d1,d2,d3):
        driver  = self.driver;
        textbox = driver.find_element_by_name("txtregno");
        textbox.clear();
        degree  = driver.find_element_by_name("cmbdegree");
        exam    = driver.find_element_by_id("cmbexamno");
        textbox.send_keys(d1);
        degree.send_keys(d2);
        exam.send_keys(d3);
        submit  = driver.find_element_by_name("button1");
        submit.click();
    def findTableHeading(self,XPATH):
        driver = self.driver;
        try:
           table = driver.find_element_by_xpath(XPATH);
           all_rows = table.find_elements_by_tag_name("tr");
           cols_heading = all_rows[0];
           t_column_heading = []
           cells = cols_heading.find_elements_by_tag_name("td");
           for cell in cells:
                t_column_heading.append(cell.text);
           self.head = t_column_heading;
        except:
           print("Mark Table doen't exist");
    def findTableBody(self,XPATH):
        driver = self.driver;
        try:
            table = driver.find_element_by_xpath(XPATH);
            all_rows = table.find_elements_by_tag_name("tr");
            subject_rows = [];
            print("all_rows  ",all_rows);
            for row in all_rows:
                cells = row.find_elements_by_tag_name("td");
                l1=[];
                for cell in cells:
                   l1.append(cell.text);
                subject_rows.append(l1);
            print(subject_rows);
            self.mark = subject_rows;
        except:
            print("Mark doesn't exist");
    def findNewHeader(self,mark):
      new_header,add_header = [],[];
      try:
        if "Subject Name" in mark[0]:
           add_header = mark[0][2:len(mark[0])]
        for row in mark:
           if row is not mark[0]:
              TornodoFetch.new_header = TornodoFetch.new_header+[row[1]]+add_header;
      except TypeError:
        pass;
    def __del__(self):
        self.driver.close();
        self.driver.quit();
class Partition(threading.Thread):
    url,markTableXPATH,degree,year,startno,endno,no,BlockNumber,completeRangeList = None,None,None,None,None,None,None,1,[];
    @staticmethod
    def globalassign(url,markTableXPATH,degree,year,startno,endno):
        Partition.url = url;
        Partition.markTableXPATH = markTableXPATH;
        Partition.degree = degree;
        Partition.year = year;
        Partition.startno = startno;
        Partition.endno = endno;
        s1=Partition.startno;
        s2=Partition.endno;
        temp="";
        dummy_set = [];
        for c in s1:
            dummy_set += c;
        for i in range(0,len(s1)):
            if (s1[i] == s2[i]):
                temp+=s1[i];
                continue;
            s1 = s1[i:len(s1)];
            s2 = s2[i:len(s2)];
            break;
        try:
          i1 = int(s1);
          i2 = int(s2);
          for i in range(i1,i2+1):
             k = str(i).zfill(len(s1));
             Partition.completeRangeList.append(temp+k);
          #print(Partition.completeRangeList);
        except Exception:
           Partition.completeRangeList = [Partition.startno];
    def __init__(self,url,degree,year,b,markTableXPATH):
        threading.Thread.__init__(self);
        self.url = url;
        self.degree = degree;
        self.year = year;
        self.markTableXPATH = markTableXPATH;
        self.BlockNumber = b;
    def run(self):
        print(self.completeRangeList);
        k = 1;
        for i in self.completeRangeList:
            fu = TornodoFetch();
            fu.URL(self.url);
            fu.insertData(i,self.degree,self.year);
            fu.findTableBody(self.markTableXPATH);
            if len(fu.mark) == 0:
                print(i," is invalid number");
                del fu;
                continue;
            print(i,"Number is valid ");
            print(k+self.BlockNumber," ",1);
            XlsxVar.ws.cell(row=k+self.BlockNumber,column=1).value=i;
            XlsxVar.save();
            j=2;
            flag1 = True;
            flag2 = True;
            for row in fu.mark:
                if(flag1):
                   flag1 = False;
                   continue;
                for col in row:
                    if (flag2):
                       flag2 = False;
                       continue;
                    #print(self.scalevalue);
                    print(k+self.BlockNumber," ",j);
                    print(col);
                    XlsxVar.ws.cell(row=k+self.BlockNumber,column=j).value=col;
                    XlsxVar.save();
                    j=j+1;
                flag2=True;
            k = k+1;
            del fu;
def Tornodo(no,degree,year,url,path):
    try:
       s = TornodoFetch();
       s.URL(url);
       s.insertData(no,degree,year);
       s.findTableHeading(path);
       s.findTableBody(path);
       s.findNewHeader(s.mark);
       del s;
    except:
       print("Invalid number is given unable to detect header");
       pass;
if __name__ ==  "__main__":
   #startno = "15td1260";
   #endno = "15td1300";
   #validno = "15td1221";
   print("Started Storming");
   #filen = "BTECH.xlsx";
   #sheet = "CSE4";
   url = "http://result.pondiuni.edu.in/candidate.asp";
   path = "/html/body/form/strong/font/font/table[4]";
   #degree = "BTHCS";
   #year = "Third";
   startno = input("Enter the starting number: ");
   endno = input("Enter the ending number: ");
   validno = input("Enter the valid number: ");
   degree = input("Enter the degree: ");
   year = input("Enter the year: ");
   filen = input("Enter the filename to be stored: ");
   sheet = input("Enter the sheetname to be stored: ");
   Partition.globalassign(url,path,degree,year,startno,endno);
   XlsxVar.filename = filen;
   XlsxVar.sheetname= sheet;
   Tornodo(validno, degree, year, url, path);
   XlsxVar.writeHeading();
   XlsxVar.save();
   #print(Partition.completeRangeList);
   for i in range(0,math.ceil((len(Partition.completeRangeList)+1)/30)):
         blockno = (i*30)+1;
         p = Partition(url,degree,year,blockno,path);
         #print(Partition.completeRangeList[(i*30):(i+1)*30],"\n\n");
         p.completeRangeList = Partition.completeRangeList[(i*30):(i+1)*30];
         p.BlockNumber =  blockno;
         p.start();
         i = i+1;
   #print(TornodoFetch.new_header);
   XlsxVar.close();
